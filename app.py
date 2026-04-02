from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import io
import json
import os
from datetime import datetime, timezone
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory, Response, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import Text

BASE_DIR = Path(__file__).resolve().parent

app = Flask(__name__, static_folder=str(BASE_DIR), static_url_path="")


def build_database_uri() -> str:
    database_url = os.getenv("DATABASE_URL", "").strip()

    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)

    if database_url:
        if "sslmode=" not in database_url:
            separator = "&" if "?" in database_url else "?"
            database_url = f"{database_url}{separator}sslmode=require"
        return database_url

    return f"sqlite:///{BASE_DIR / 'app.db'}"


app.config["SQLALCHEMY_DATABASE_URI"] = build_database_uri()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["JSON_SORT_KEYS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}


db = SQLAlchemy(app)


class Record(db.Model):
    __tablename__ = "records"

    id = db.Column(db.String(120), primary_key=True)
    record_type = db.Column(db.String(50), nullable=True, index=True)
    data = db.Column(Text, nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(
        db.DateTime(timezone=True),
        nullable=False,
        default=lambda: datetime.now(timezone.utc),
        onupdate=lambda: datetime.now(timezone.utc),
    )

    def to_dict(self) -> dict:
        payload = json.loads(self.data)
        if not isinstance(payload, dict):
            raise ValueError("Stored record is not a JSON object")
        if "id" not in payload:
            payload["id"] = self.id
        if "type" not in payload and self.record_type:
            payload["type"] = self.record_type
        return payload


class AccessLog(db.Model):
    __tablename__ = "access_logs"

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(120), nullable=False, index=True)
    action = db.Column(db.String(50), nullable=False, index=True)
    result = db.Column(db.String(30), nullable=False, index=True)
    source = db.Column(db.String(50), nullable=False, default="panes")
    ip_address = db.Column(db.String(120), nullable=True)
    user_agent = db.Column(Text, nullable=True)
    details = db.Column(Text, nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=lambda: datetime.now(timezone.utc), index=True)

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "username": self.username,
            "action": self.action,
            "result": self.result,
            "source": self.source,
            "ip_address": self.ip_address,
            "user_agent": self.user_agent,
            "details": self.details,
            "created_at": self.created_at.astimezone(timezone.utc).isoformat() if self.created_at else None,
        }


class AuditLog(db.Model):
    __tablename__ = "audit_logs"

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(120), nullable=False, index=True)
    action = db.Column(db.String(50), nullable=False, index=True)
    source = db.Column(db.String(50), nullable=False, default="panes", index=True)
    target_type = db.Column(db.String(80), nullable=True, index=True)
    target_id = db.Column(db.String(120), nullable=True, index=True)
    details = db.Column(Text, nullable=True)
    ip_address = db.Column(db.String(120), nullable=True)
    user_agent = db.Column(Text, nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=lambda: datetime.now(timezone.utc), index=True)

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "username": self.username,
            "action": self.action,
            "source": self.source,
            "target_type": self.target_type,
            "target_id": self.target_id,
            "details": self.details,
            "ip_address": self.ip_address,
            "user_agent": self.user_agent,
            "created_at": self.created_at.astimezone(timezone.utc).isoformat() if self.created_at else None,
        }


with app.app_context():
    db.create_all()


def parse_iso_datetime(value):
    if not value:
        return None

    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return None
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        try:
            dt = datetime.fromisoformat(text)
        except ValueError:
            return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def normalize_tipo(value: str) -> str:
    text = str(value or "").strip().lower()
    if text in {"avionico", "aviônico"}:
        return "avionico"
    if text == "mecânico":
        return "mecanico"
    return text


def pane_status_aberta(status: str) -> bool:
    return normalize_tipo(status) != "finalizada" and str(status or "").strip().lower() != "finalizada"


def montar_ranking_panes(tipo: str | None = None, limit: int = 10):
    records = Record.query.order_by(Record.created_at.asc(), Record.id.asc()).all()

    aeronaves = {}
    panes = []

    status_label_map = {
        "lancada": "Monitoradas",
        "em_progresso": "Em Progresso",
        "aguardando_material": "Aguardando Material",
        "aguardando_ferramenta": "Aguardando Ferramenta",
        "aguardando_transferencia": "Aguardando Transferência",
    }

    status_order_map = {
        "lancada": 1,
        "em_progresso": 2,
        "aguardando_material": 3,
        "aguardando_ferramenta": 4,
        "aguardando_transferencia": 5,
    }

    for record in records:
        try:
            payload = record.to_dict()
        except Exception:
            continue

        record_type = str(payload.get("type") or record.record_type or "").strip().lower()

        if record_type == "aeronave":
            aeronaves[str(payload.get("id", "")).strip()] = payload
            continue

        if record_type != "pane":
            continue

        pane_tipo = normalize_tipo(payload.get("paneTipo"))
        pane_status = str(payload.get("paneStatus", "")).strip().lower()

        if tipo and pane_tipo != normalize_tipo(tipo):
            continue

        if not pane_status_aberta(pane_status):
            continue

        criado_em = parse_iso_datetime(payload.get("criadoEm")) or record.created_at
        if not criado_em:
            continue

        agora = datetime.now(timezone.utc)
        delta = agora - criado_em
        total_seconds = max(0, delta.total_seconds())
        horas = round(total_seconds / 3600, 1)
        dias = max(1, int(total_seconds // 86400))

        if dias <= 7:
            sla_status = "dentro_sla"
            sla_label = "Dentro do SLA"
            antiguidade_faixa = "normal"
        elif dias <= 14:
            sla_status = "atencao"
            sla_label = "Atenção SLA"
            antiguidade_faixa = "atencao"
        else:
            sla_status = "vencido"
            sla_label = "SLA Vencido"
            antiguidade_faixa = "critico"

        aeronave_id = str(payload.get("aeronaveId", "")).strip()
        aeronave = aeronaves.get(aeronave_id, {})

        panes.append(
            {
                "id": str(payload.get("id", "")).strip(),
                "aeronaveId": aeronave_id,
                "prefixo": aeronave.get("prefixo"),
                "modelo": aeronave.get("modelo"),
                "descricao": payload.get("paneDescricao", ""),
                "ata": payload.get("paneAta"),
                "tipo": pane_tipo,
                "status": pane_status,
                "statusLabel": status_label_map.get(pane_status, pane_status or "Sem status"),
                "statusOrder": status_order_map.get(pane_status, 999),
                "criadoEm": criado_em.isoformat(),
                "horasEmAberto": horas,
                "diasEmAberto": dias,
                "diasTexto": f"{dias} dia{'s' if dias != 1 else ''} pendente",
                "slaStatus": sla_status,
                "slaLabel": sla_label,
                "antiguidadeFaixa": antiguidade_faixa,
                "criadoPor": payload.get("criadoPor"),
            }
        )

    panes.sort(
        key=lambda item: (
            -int(item.get("diasEmAberto") or 0),
            -float(item.get("horasEmAberto") or 0),
            int(item.get("statusOrder") or 999),
            str(item.get("prefixo") or ""),
            str(item.get("descricao") or ""),
        )
    )
    return panes[:limit]




@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin", "*")
    response.headers["Access-Control-Allow-Origin"] = origin if origin else "*"
    response.headers["Vary"] = "Origin"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response


@app.route("/api/<path:_>", methods=["OPTIONS"])
@app.route("/<path:_>", methods=["OPTIONS"])
def options_handler(_):
    return ("", 204)


def get_client_ip() -> str | None:
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip() or None
    real_ip = request.headers.get("X-Real-IP", "").strip()
    if real_ip:
        return real_ip
    return request.remote_addr


def create_access_log(username: str, action: str, result: str, source: str = "panes", details: str | None = None):
    log = AccessLog(
        username=str(username or "desconhecido")[:120],
        action=str(action or "unknown")[:50],
        result=str(result or "unknown")[:30],
        source=str(source or "panes")[:50],
        ip_address=(get_client_ip() or "")[:120] or None,
        user_agent=(request.headers.get("User-Agent", "") or "")[:2000] or None,
        details=(str(details)[:4000] if details else None),
    )
    db.session.add(log)
    db.session.commit()
    return log




def apply_access_log_filters(query):
    username = str(request.args.get("username", "")).strip()
    action = str(request.args.get("action", "")).strip()
    result = str(request.args.get("result", "")).strip()
    text = str(request.args.get("q", "")).strip()
    date_from = parse_iso_datetime(request.args.get("date_from"))
    date_to = parse_iso_datetime(request.args.get("date_to"))

    if username:
        query = query.filter(AccessLog.username.ilike(f"%{username}%"))
    if action:
        query = query.filter(AccessLog.action == action)
    if result:
        query = query.filter(AccessLog.result == result)
    if date_from:
        query = query.filter(AccessLog.created_at >= date_from)
    if date_to:
        query = query.filter(AccessLog.created_at <= date_to)

    if text:
        from sqlalchemy import or_
        query = query.filter(
            or_(
                AccessLog.username.ilike(f"%{text}%"),
                AccessLog.details.ilike(f"%{text}%"),
                AccessLog.ip_address.ilike(f"%{text}%"),
            )
        )

    return query


def build_csv_response(rows, filename: str):
    output = io.StringIO()
    if rows:
        fieldnames = list(rows[0].keys())
    else:
        fieldnames = ["message"]
        rows = [{"message": "Sem registros"}]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

    return Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



def infer_actor_username(payload: dict | None = None) -> str:
    header_username = str(request.headers.get("X-Actor-Username", "")).strip()
    if header_username:
        return header_username[:120]

    if isinstance(payload, dict):
        for key in ("modified_by", "created_by", "criadoPor", "updatedBy", "createdBy", "username", "responsavel", "login_usuario"):
            value = str(payload.get(key, "")).strip()
            if value:
                return value[:120]

    return "desconhecido"


def create_audit_log(username: str, action: str, source: str = "panes", target_type: str | None = None, target_id: str | None = None, details: str | None = None):
    log = AuditLog(
        username=str(username or "desconhecido")[:120],
        action=str(action or "unknown")[:50],
        source=str(source or "panes")[:50],
        target_type=(str(target_type)[:80] if target_type else None),
        target_id=(str(target_id)[:120] if target_id else None),
        details=(str(details)[:4000] if details else None),
        ip_address=(get_client_ip() or "")[:120] or None,
        user_agent=(request.headers.get("User-Agent", "") or "")[:2000] or None,
    )
    db.session.add(log)
    db.session.commit()
    return log


def apply_audit_log_filters(query):
    username = str(request.args.get("username", "")).strip()
    action = str(request.args.get("action", "")).strip()
    source = str(request.args.get("source", "")).strip()
    text = str(request.args.get("q", "")).strip()
    date_from = parse_iso_datetime(request.args.get("date_from"))
    date_to = parse_iso_datetime(request.args.get("date_to"))

    if username:
        query = query.filter(AuditLog.username.ilike(f"%{username}%"))
    if action:
        query = query.filter(AuditLog.action == action)
    if source:
        query = query.filter(AuditLog.source == source)
    if date_from:
        query = query.filter(AuditLog.created_at >= date_from)
    if date_to:
        query = query.filter(AuditLog.created_at <= date_to)

    if text:
        from sqlalchemy import or_
        query = query.filter(
            or_(
                AuditLog.username.ilike(f"%{text}%"),
                AuditLog.details.ilike(f"%{text}%"),
                AuditLog.target_id.ilike(f"%{text}%"),
                AuditLog.target_type.ilike(f"%{text}%"),
                AuditLog.ip_address.ilike(f"%{text}%"),
            )
        )
    return query


def friendly_audit_row(row: dict) -> dict:
    action_map = {
        "create": "Criou registro",
        "update": "Atualizou registro",
        "delete": "Excluiu registro",
    }
    source_map = {
        "panes": "Controle Técnico da Frota",
        "rodend": "Controle Rod End",
    }
    type_map = {
        "aeronave": "Aeronave",
        "pane": "Discrepância",
        "etapa": "Etapa",
        "pendencia": "Pendência",
        "usuario": "Usuário",
        "rodend_user": "Usuário Rod End",
        "rodend_aircraft": "Aeronave Rod End",
        "rodend_component": "Componente Rod End",
    }

    raw_created_at = row.get("created_at") or ""
    try:
        created_label = datetime.fromisoformat(str(raw_created_at).replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        created_label = str(raw_created_at or "")

    return {
        "Data/Hora": created_label,
        "Usuário": row.get("username", ""),
        "Ação": action_map.get(str(row.get("action", "")).strip(), str(row.get("action", "")).strip()),
        "Módulo": source_map.get(str(row.get("source", "")).strip(), str(row.get("source", "")).strip()),
        "Tipo de registro": type_map.get(str(row.get("target_type", "")).strip(), str(row.get("target_type", "")).strip()),
        "ID do registro": row.get("target_id", ""),
        "Detalhes": row.get("details", ""),
        "IP": row.get("ip_address", ""),
    }

def build_excel_response(rows, filename: str, sheet_name: str = "Auditoria"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    if rows:
        headers = list(rows[0].keys())
    else:
        headers = ["message"]
        rows = [{"message": "Sem registros"}]

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in sheet.columns:
        length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                length = max(length, len(str(cell.value or "")))
            except Exception:
                pass
        sheet.column_dimensions[column].width = min(max(length + 2, 12), 40)

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/health")
def health():
    return jsonify({"ok": True, "service": "controle-panes-api"})


@app.post("/api/access-log")
def create_access_log_endpoint():
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"error": "JSON inválido"}), 400

    username = str(payload.get("username", "")).strip() or "desconhecido"
    action = str(payload.get("action", "")).strip() or "unknown"
    result = str(payload.get("result", "")).strip() or "unknown"
    source = str(payload.get("source", "panes")).strip() or "panes"
    details = str(payload.get("details", "")).strip() or None

    log = create_access_log(username=username, action=action, result=result, source=source, details=details)
    return jsonify({"ok": True, "item": log.to_dict()}), 201


@app.get("/api/access-log")
def list_access_logs():
    limit = request.args.get("limit", default=200, type=int)
    if limit is None or limit < 1:
        limit = 200
    if limit > 1000:
        limit = 1000

    query = AccessLog.query.order_by(AccessLog.created_at.desc(), AccessLog.id.desc())
    query = apply_access_log_filters(query)
    logs = query.limit(limit).all()
    return jsonify([log.to_dict() for log in logs])


@app.get("/api/access-log/export")
def export_access_logs():
    query = AccessLog.query.order_by(AccessLog.created_at.desc(), AccessLog.id.desc())
    query = apply_access_log_filters(query)
    logs = query.limit(5000).all()
    return build_csv_response([log.to_dict() for log in logs], "auditoria_acesso.csv")




@app.get("/api/access-log/summary")
def access_log_summary():
    query = AccessLog.query
    query = apply_access_log_filters(query)
    logs = query.order_by(AccessLog.created_at.desc(), AccessLog.id.desc()).limit(5000).all()

    today = datetime.now(timezone.utc).date()
    login_today = 0
    fail_today = 0
    user_counts = {}
    latest_by_user = {}
    ip_counts = {}
    fail_by_user = {}
    fail_by_ip = {}
    last_success_by_user = {}

    for log in logs:
        created_at = log.created_at.astimezone(timezone.utc) if log.created_at else None
        if created_at and created_at.date() == today:
            if log.action == "login" and log.result == "success":
                login_today += 1
            if log.result == "fail":
                fail_today += 1

        username = str(log.username or "desconhecido").strip() or "desconhecido"
        user_counts[username] = user_counts.get(username, 0) + 1

        if username not in latest_by_user:
            latest_by_user[username] = {
                "username": username,
                "action": log.action,
                "result": log.result,
                "created_at": created_at.isoformat() if created_at else None,
                "ip_address": log.ip_address,
            }

        if log.result == "success" and username not in last_success_by_user:
            last_success_by_user[username] = {
                "username": username,
                "created_at": created_at.isoformat() if created_at else None,
                "ip_address": log.ip_address,
            }

        ip = str(log.ip_address or "").strip()
        if ip:
            ip_counts[ip] = ip_counts.get(ip, 0) + 1

        if log.result == "fail":
            fail_by_user[username] = fail_by_user.get(username, 0) + 1
            if ip:
                fail_by_ip[ip] = fail_by_ip.get(ip, 0) + 1

    top_users = [
        {"username": username, "count": count}
        for username, count in sorted(user_counts.items(), key=lambda item: (-item[1], item[0]))[:5]
    ]

    latest_accesses = list(latest_by_user.values())[:8]

    top_ips = [
        {"ip_address": ip, "count": count}
        for ip, count in sorted(ip_counts.items(), key=lambda item: (-item[1], item[0]))[:5]
    ]

    top_fail_users = [
        {"username": username, "count": count}
        for username, count in sorted(fail_by_user.items(), key=lambda item: (-item[1], item[0]))[:5]
    ]

    top_fail_ips = [
        {"ip_address": ip, "count": count}
        for ip, count in sorted(fail_by_ip.items(), key=lambda item: (-item[1], item[0]))[:8]
    ]

    suspicious_ips = []
    for ip, count in sorted(fail_by_ip.items(), key=lambda item: (-item[1], item[0])):
        severity = "high" if count >= 10 else ("medium" if count >= 5 else "low")
        if severity in {"high", "medium"}:
            suspicious_ips.append({
                "ip_address": ip,
                "fail_count": count,
                "severity": severity,
            })

    last_success_list = list(last_success_by_user.values())[:8]

    return jsonify({
        "login_today": login_today,
        "fail_today": fail_today,
        "top_users": top_users,
        "latest_accesses": latest_accesses,
        "top_ips": top_ips,
        "top_fail_users": top_fail_users,
        "top_fail_ips": top_fail_ips,
        "last_success_by_user": last_success_list,
        "suspicious_ips": suspicious_ips,
        "total_considered": len(logs),
    })



@app.get("/api/audit-log")
def list_audit_logs():
    limit = request.args.get("limit", default=300, type=int)
    if limit is None or limit < 1:
        limit = 300
    if limit > 2000:
        limit = 2000

    query = AuditLog.query.order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
    query = apply_audit_log_filters(query)
    logs = query.limit(limit).all()
    return jsonify([log.to_dict() for log in logs])


@app.get("/api/audit-log/export")
def export_audit_logs():
    query = AuditLog.query.order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
    query = apply_audit_log_filters(query)
    logs = query.limit(5000).all()
    return build_excel_response([friendly_audit_row(log.to_dict()) for log in logs], "auditoria_acoes_formatada.xlsx", "Auditoria de Ações")

@app.get("/api/records")
def list_records():
    records = Record.query.order_by(Record.created_at.asc(), Record.id.asc()).all()
    return jsonify([record.to_dict() for record in records])


@app.post("/api/records")
def create_record():
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"error": "JSON inválido"}), 400

    record_id = str(payload.get("id", "")).strip()
    if not record_id:
        return jsonify({"error": "Campo 'id' é obrigatório"}), 400

    existing = db.session.get(Record, record_id)
    if existing is not None:
        return jsonify({"error": "Já existe um registro com este id"}), 409

    record = Record(
        id=record_id,
        record_type=str(payload.get("type", "")).strip() or None,
        data=json.dumps(payload, ensure_ascii=False),
    )
    db.session.add(record)
    db.session.commit()

    actor = infer_actor_username(payload)
    record_type = str(payload.get("type", "")).strip() or "registro"
    create_audit_log(
        username=actor,
        action="create",
        source="panes",
        target_type=record_type,
        target_id=record_id,
        details=f"Criação de registro tipo={record_type} id={record_id}",
    )
    return jsonify({"isOk": True, "item": record.to_dict()}), 201


@app.put("/api/records/<record_id>")
def update_record(record_id: str):
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"error": "JSON inválido"}), 400

    record = db.session.get(Record, record_id)
    if record is None:
        return jsonify({"error": "Registro não encontrado"}), 404

    payload["id"] = record_id
    record.record_type = str(payload.get("type", "")).strip() or None
    record.data = json.dumps(payload, ensure_ascii=False)
    record.updated_at = datetime.now(timezone.utc)
    db.session.commit()

    actor = infer_actor_username(payload)
    record_type = str(payload.get("type", "")).strip() or "registro"
    create_audit_log(
        username=actor,
        action="update",
        source="panes",
        target_type=record_type,
        target_id=record_id,
        details=f"Atualização de registro tipo={record_type} id={record_id}",
    )
    return jsonify({"isOk": True, "item": record.to_dict()})


@app.delete("/api/records/<record_id>")
def delete_record(record_id: str):
    record = db.session.get(Record, record_id)
    if record is None:
        return jsonify({"error": "Registro não encontrado"}), 404

    payload = None
    try:
        payload = record.to_dict()
    except Exception:
        payload = {}

    actor = infer_actor_username(payload)
    target_type = str(payload.get("type") or record.record_type or "registro").strip() or "registro"
    db.session.delete(record)
    db.session.commit()
    create_audit_log(
        username=actor,
        action="delete",
        source="panes",
        target_type=target_type,
        target_id=record_id,
        details=f"Exclusão de registro tipo={target_type} id={record_id}",
    )
    return jsonify({"isOk": True, "deletedId": record_id})


@app.delete("/api/records")
def delete_record_from_body():
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"error": "JSON inválido"}), 400

    record_id = str(payload.get("id", "")).strip()
    if not record_id:
        return jsonify({"error": "Campo 'id' é obrigatório"}), 400
    return delete_record(record_id)




@app.get("/api/ranking-panes")
def ranking_panes():
    tipo = request.args.get("tipo")
    limit = request.args.get("limit", default=10, type=int)

    if limit is None or limit < 1:
        limit = 10
    if limit > 50:
        limit = 50

    ranking = montar_ranking_panes(tipo=tipo, limit=limit)
    return jsonify(ranking)

@app.get("/")
def index():
    candidates = [
        "codigo_render_api.html",
        "index.html",
        "codigo completo.html",
    ]
    for filename in candidates:
        file_path = BASE_DIR / filename
        if file_path.exists():
            return send_from_directory(BASE_DIR, filename)
    return jsonify(
        {
            "ok": True,
            "message": "API online. Coloque seu HTML no mesmo diretório do app.py com o nome 'codigo_render_api.html' ou 'index.html'.",
        }
    )


# =========================
# ROD END - HELPERS
# =========================

RODEND_ALLOWED_TYPES = {"rodend_user", "rodend_aircraft", "rodend_component"}


def is_rodend_type(value: str) -> bool:
    return str(value or "").strip().lower() in RODEND_ALLOWED_TYPES


def get_rodend_records():
    records = Record.query.order_by(Record.created_at.asc(), Record.id.asc()).all()
    items = []

    for record in records:
        try:
            payload = record.to_dict()
        except Exception:
            continue

        record_type = str(payload.get("type") or record.record_type or "").strip().lower()
        if is_rodend_type(record_type):
            items.append(payload)

    return items


def get_rodend_users():
    return [item for item in get_rodend_records() if str(item.get("type", "")).strip().lower() == "rodend_user"]


def sanitize_rodend_payload(payload: dict, record_id: str | None = None):
    if not isinstance(payload, dict):
        return None, ("JSON inválido", 400)

    payload_type = str(payload.get("type", "")).strip().lower()
    if not is_rodend_type(payload_type):
        return None, ("Tipo inválido para módulo ROD END", 400)

    if record_id:
        payload["id"] = record_id

    final_id = str(payload.get("id", "")).strip()
    if not final_id:
        return None, ("Campo 'id' é obrigatório", 400)

    # Reforça prefixos para evitar colisão com o sistema atual
    if payload_type == "rodend_user" and not final_id.startswith("roduser_"):
        return None, ("ID de usuário ROD END deve começar com 'roduser_'", 400)

    if payload_type == "rodend_aircraft" and not final_id.startswith("rodair_"):
        return None, ("ID de aeronave ROD END deve começar com 'rodair_'", 400)

    if payload_type == "rodend_component" and not final_id.startswith("rodcomp_"):
        return None, ("ID de componente ROD END deve começar com 'rodcomp_'", 400)

    payload["id"] = final_id
    payload["type"] = payload_type

    return payload, None


# =========================
# ROD END - ROTAS HTML
# =========================

@app.get("/rodend")
def rodend_index():
    candidates = [
        "rodend.html",
        "Rod End Control Ultima Versao 24-03-2026.html",
    ]
    for filename in candidates:
        file_path = BASE_DIR / filename
        if file_path.exists():
            return send_from_directory(BASE_DIR, filename)

    return jsonify(
        {
            "ok": False,
            "message": "Arquivo do módulo ROD END não encontrado. Use 'rodend.html' no mesmo diretório do app.py.",
        }
    ), 404


# =========================
# ROD END - API
# =========================

@app.post("/api/rodend/login")
def rodend_login():
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return jsonify({"ok": False, "error": "JSON inválido"}), 400

    username = str(payload.get("username", "")).strip()
    password = str(payload.get("password", "")).strip()

    if not username or not password:
        return jsonify({"ok": False, "error": "Usuário e senha são obrigatórios"}), 400

    # admin fixo do módulo ROD END
    if username == "admin" and password == "Omni0320!":
        return jsonify(
            {
                "ok": True,
                "user": {
                    "id": "roduser_admin",
                    "username": "admin",
                    "type": "rodend_user",
                    "isAdmin": True,
                    "needsPasswordChange": False,
                }
            }
        )

    users = get_rodend_users()
    user = next((u for u in users if str(u.get("username", "")).strip() == username), None)

    if not user or str(user.get("password", "")).strip() != password:
        return jsonify({"ok": False, "error": "Usuário ou senha incorretos"}), 401

    return jsonify(
        {
            "ok": True,
            "user": {
                "id": user.get("id"),
                "username": user.get("username"),
                "type": "rodend_user",
                "isAdmin": False,
                "needsPasswordChange": bool(user.get("needs_password_change", False)),
            }
        }
    )


@app.get("/api/rodend/records")
def rodend_list_records():
    return jsonify(get_rodend_records())


@app.post("/api/rodend/records")
def rodend_create_record():
    payload = request.get_json(silent=True)
    payload, error = sanitize_rodend_payload(payload)

    if error:
        return jsonify({"error": error[0]}), error[1]

    existing = db.session.get(Record, payload["id"])
    if existing is not None:
        return jsonify({"error": "Já existe um registro com este id"}), 409

    record = Record(
        id=payload["id"],
        record_type=payload["type"],
        data=json.dumps(payload, ensure_ascii=False),
    )
    db.session.add(record)
    db.session.commit()

    actor = infer_actor_username(payload)
    create_audit_log(
        username=actor,
        action="create",
        source="rodend",
        target_type=payload.get("type"),
        target_id=payload.get("id"),
        details=f"Criação de registro ROD END tipo={payload.get('type')} id={payload.get('id')}",
    )
    return jsonify({"isOk": True, "item": record.to_dict()}), 201


@app.put("/api/rodend/records/<record_id>")
def rodend_update_record(record_id: str):
    payload = request.get_json(silent=True)
    payload, error = sanitize_rodend_payload(payload, record_id=record_id)

    if error:
        return jsonify({"error": error[0]}), error[1]

    record = db.session.get(Record, record_id)
    if record is None:
        return jsonify({"error": "Registro não encontrado"}), 404

    current_type = str(record.record_type or "").strip().lower()
    if not is_rodend_type(current_type):
        return jsonify({"error": "Registro não pertence ao módulo ROD END"}), 400

    record.record_type = payload["type"]
    record.data = json.dumps(payload, ensure_ascii=False)
    record.updated_at = datetime.now(timezone.utc)
    db.session.commit()

    actor = infer_actor_username(payload)
    create_audit_log(
        username=actor,
        action="update",
        source="rodend",
        target_type=payload.get("type"),
        target_id=payload.get("id"),
        details=f"Atualização de registro ROD END tipo={payload.get('type')} id={payload.get('id')}",
    )
    return jsonify({"isOk": True, "item": record.to_dict()})


@app.delete("/api/rodend/records/<record_id>")
def rodend_delete_record(record_id: str):
    record = db.session.get(Record, record_id)
    if record is None:
        return jsonify({"error": "Registro não encontrado"}), 404

    current_type = str(record.record_type or "").strip().lower()
    if not is_rodend_type(current_type):
        return jsonify({"error": "Registro não pertence ao módulo ROD END"}), 400

    payload = None
    try:
        payload = record.to_dict()
    except Exception:
        payload = {}

    actor = infer_actor_username(payload)
    target_type = str(payload.get("type") or record.record_type or "rodend_registro").strip() or "rodend_registro"
    db.session.delete(record)
    db.session.commit()
    create_audit_log(
        username=actor,
        action="delete",
        source="rodend",
        target_type=target_type,
        target_id=record_id,
        details=f"Exclusão de registro ROD END tipo={target_type} id={record_id}",
    )
    return jsonify({"isOk": True, "deletedId": record_id})

@app.get("/<path:filename>")
def serve_static(filename: str):
    file_path = BASE_DIR / filename
    if file_path.exists() and file_path.is_file():
        return send_from_directory(BASE_DIR, filename)
    return jsonify({"error": "Arquivo não encontrado"}), 404

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False)
